#!/usr/bin/env python
import re
import logging
import click
from pathlib import Path
from openpyxl import load_workbook

from parsers.psgc.read import read
from parsers.psgc.transform import (
    transform_main,
    transform_extended,
    transform_flat,
    transform_fuzzer,
)
from parsers.psgc.export import export

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def score_excel_file(file_path: Path) -> int:
    """Score an Excel file based on likelihood of containing PSGC data."""
    name = file_path.name.lower()
    score = 0
    if "publication" in name:
        score += 5
    if "psgc" in name:
        score += 3
    return score


def find_psgc_file(dir_path: Path) -> Path | None:
    """Find the Excel file containing PSGC sheet in a directory."""
    excel_files = [f for f in dir_path.glob("*.xlsx") if f.is_file()]
    if not excel_files:
        return None

    # Score files and sort
    scored = [(score_excel_file(f), f.stat().st_size, f) for f in excel_files]
    scored.sort(key=lambda x: (-x[0], -x[1]))

    # Try files by score until one has PSGC sheet
    for _, _, file_path in scored:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            if "PSGC" in wb.sheetnames:
                return file_path
        except Exception:
            continue
    return None


def process_directory(input_dir: Path, output_dir: Path) -> bool:
    """Process a single directory (date folder). Returns True on success, False on failure."""
    excel_file = find_psgc_file(input_dir)
    if not excel_file:
        logger.warning(f"No valid PSGC file found in {input_dir}")
        return False

    click.echo(f"  Processing {excel_file.name}")
    try:
        df = read(excel_file)

        # Transform
        main_data = transform_main(df)
        extended_data = transform_extended(df)
        flat_data = transform_flat(df)
        fuzzer_base = transform_fuzzer(df)

        # Export
        export(output_dir, main_data)
        export(output_dir, extended_data)
        export(output_dir, flat_data)
        export(output_dir, fuzzer_base)

        return True
    except Exception as e:
        logger.error(f"Failed to process {input_dir}: {e}")
        return False


@click.command()
@click.argument("input_dir", type=click.Path(exists=True, path_type=Path))
@click.argument("output_dir", type=click.Path(path_type=Path))
def barpar(input_dir: Path, output_dir: Path) -> None:
    """Process PSGC Excel files from INPUT_DIR to OUTPUT_DIR."""
    input_dir = input_dir.expanduser().resolve()
    output_dir = output_dir.expanduser().resolve()

    # Check if input is a single file
    if input_dir.is_file():
        click.echo(f"Processing single file: {input_dir}")
        if not input_dir.suffix == ".xlsx":
            click.echo("Error: Input file must be an Excel file")
            return

        # Use input file's parent directory name as date
        date_str = input_dir.parent.name
        if not DATE_PATTERN.match(date_str):
            click.echo(
                f"Warning: Parent directory '{date_str}' does not match YYYY-MM-DD format"
            )
            date_str = "current"

        output_path = output_dir / date_str
        output_path.mkdir(parents=True, exist_ok=True)

        # Read and process
        try:
            df = read(input_dir)
            main_data = transform_main(df)
            extended_data = transform_extended(df)
            flat_data = transform_flat(df)
            fuzzer_base = transform_fuzzer(df)
            export(output_path, main_data)
            export(output_path, extended_data)
            export(output_path, flat_data)
            export(output_path, fuzzer_base)

            click.echo(f"Output written to {output_path}")
        except Exception as e:
            logger.error(f"Failed to process {input_dir}: {e}")
        return

    # Check if input is a single directory (non-date)
    if input_dir.is_dir():
        subdirs = [d for d in input_dir.iterdir() if d.is_dir()]
        if not subdirs:
            # Single directory with files
            click.echo(f"Processing single directory: {input_dir}")
            excel_file = find_psgc_file(input_dir)
            if not excel_file:
                logger.error(f"No valid PSGC file found in {input_dir}")
                return

            date_str = input_dir.name
            if not DATE_PATTERN.match(date_str):
                date_str = "current"

            output_path = output_dir / date_str
            output_path.mkdir(parents=True, exist_ok=True)

            try:
                df = read(excel_file)
                main_data = transform_main(df)
                extended_data = transform_extended(df)
                flat_data = transform_flat(df)
                fuzzer_base = transform_fuzzer(df)
                export(output_path, main_data)
                export(output_path, extended_data)
                export(output_path, flat_data)
                export(output_path, fuzzer_base)

                click.echo(f"Output written to {output_path}")
            except Exception as e:
                logger.error(f"Failed to process {input_dir}: {e}")
            return

        # Batch mode: process dated subdirectories
        dated_dirs = [d for d in subdirs if DATE_PATTERN.match(d.name)]
        if not dated_dirs:
            click.echo("Error: No dated directories (YYYY-MM-DD) found in input")
            return

        dated_dirs.sort()
        click.echo(f"Found {len(dated_dirs)} dated directories")

        success_count = 0
        for date_dir in dated_dirs:
            click.echo(f"\nProcessing {date_dir.name}")
            output_path = output_dir / date_dir.name
            output_path.mkdir(parents=True, exist_ok=True)
            if process_directory(date_dir, output_path):
                success_count += 1

        click.echo(
            f"\nProcessed {success_count}/{len(dated_dirs)} directories successfully"
        )


if __name__ == "__main__":
    barpar()
